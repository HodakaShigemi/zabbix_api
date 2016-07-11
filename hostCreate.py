from pyzabbix import ZabbixAPI
from openpyxl import load_workbook
import csv

class HostImport(object):
    """Import informations about hosts into ZabbixServer
    This module intend to make it easy that importing various information to zabbix-server.
    This module has following functions.
        - create host information
        - update host information
        - create or update host information by case that the host presents or not.

    """
    def __init__(self, server='http://localhost/zabbix', user = 'admin', password = 'zabbix'):
        self.server = server
        self.user = user
        self.password = password
        self.zapi = ZabbixAPI(server = self.server)
        self.zapi.login(user = self.user, password = self.password)
        self.zabbixHostGroups = {group["name"]:group["groupid"] for group in self.zapi.hostgroup.get()}

    def hostPresent(self, **kwargs):
        hosts = self.zapi.host.get(output = ['host'], filter = kwargs)
        if len(hosts) == 0:
            return False
        elif len(hosts) ==  1:
            return hosts[0]['hostid']
        else:
            return hosts

    def hostGroupsUpdate(self, groups):
        for groupName in groups:
            if groupName in self.zabbixHostGroups:
                pass
            else:
                result = self.zapi.hostgroup.create(name = groupName)
                self.zabbixHostGroups[groupName] = result["groupids"][0]

    def registerHost(self, host, name, groups, ip, protocol = 'snmp', inventory= {}, macros = []):
        self.hostGroupsUpdate(groups)
        if self.hostPresent(host = host):
            raise HostAlreadyExistsError('${host} is already registerd.'.format(host = host))
        else:
            protocol = protocol.lower()
            if protocol == "snmp":
                interfaceType = 2
                portNum = "161"
            elif protocol == "agent":
                interfaceType = 1
                portNum = "10050"
            else:
                raise UnknownProtocolError("${protocol} is unsupported protocol in this module.".format(protocol=protocol))
            belongingGroups = []
            for group in groups:
                belongingGroups.append({'groupid':self.zabbixHostGroups[group]})
            return self.zapi.host.create(
                host = host,
                name = name,
                interfaces = [{"type":interfaceType, "main":1, "useip":1, "ip":ip, "dns":"", "port":portNum}],
                groups = belongingGroups,
                inventory_mode = 0,
                inventory = inventory,
                macros = macros
                )

    def updateHost(self, host, name, groups, protocol = 'snmp', inventory= {}, macros = []):
        self.hostGroupsUpdate(groups)
        belongingGroups = []
        for group in groups:
            belongingGroups.append({'groupid' : self.zabbixHostGroups[group]})
        hostid = self.hostPresent(host=host)
        return self.zapi.host.update(
            hostid = hostid,
            name =name,
            host=host,
            groups = belongingGroups,
            inventory_mode = 0,
            inventory = inventory,
            macros = macros
            )

    def registerHostsFromExcel(self, pathOfExcel, rowOfItemName, sheetName):
        wb = load_workbook(pathOfExcel)
        ws = wb.get_sheet_by_name(sheetName)
        itemIndexes = {}
        for columnIndex in range(1, ws.max_column):
            itemName = ws.cell(row = rowOfItemName, column = columnIndex).value
            if itemName != None:
                itemIndexes[itemName] = columnIndex

        for row in ws.rows[rowOfItemName:]:
            for itemName in itemIndexes:
                if row[itemIndexes[itemName]].value != None:
                    itemValueOfEachHost[itemName] = row[itemIndexes[itemName]].value
            groups = []
            for key in itemValueOfEachHost:
                if 'group' in key:
                    groups.append(itemValueOfEachHost[key])
            itemValueOfEachHost["groups"] = groups
            self.registerHost(**itemValueOfEachHost)

    def registerHostsFromCsv(self, pathOfCsv):
        with open(pathOfCsv) as csvFile:
            reader = csv.DictReader(csvFile)
            for row in reader:
                procurement_macro, procurement_group = row['区分_macro'], row['区分_group']
                dc_macro, dc_group = row['DC_macro'], row['DC_group']
                kk = row['所属ｸﾞﾙｰﾌﾟ']
                gmc = 'GMC_General'
                name, ipAddr = row['名前'], row['IPｱﾄﾞﾚｽ(ﾎｽﾄ名)']
                chassis = row['機種']
                chassisAndOS = row['ｼｽﾃﾑ記述']
                site_rack = row['ｺﾒﾝﾄ1']
                construct = row['ｺﾒﾝﾄ2']
                macros = [{'macro':'{$CHOUTATSU}', 'value':procurement_macro},
                          {'macro':'{$DC}', 'value':dc_macro},
                          {'macro':'{$KK}', 'value':kk},
                          {'macro':'{$LOCATION}', 'value':site_rack},
                          {'macro':'{$KOUSEI}', 'value':construct}]
                inventory ={'name':name,
                            'chassis':chassis,
                            'model':chassisAndOS,
                            'site_rack':site_rack,
                            'site_notes':construct}
                if self.hostPresent(host = ipAddr):
                    self.updateHost(host = ipAddr,
                                    name = kk + ' ' + name,
                                    groups = [procurement_group, dc_group, kk, gmc],
                                    inventory = inventory,
                                    macros = macros)
    """
                else:
                    self.registerHost(host = ipAddr,
                                      ip = ipAddr,
                                      name = name,
                                      groups = [procurement_group, dc_group, kk, gmc],
                                      inventory = inventory,
                                      macros = macros)
    """
